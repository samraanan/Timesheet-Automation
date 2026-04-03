"""
Script to auto-generate all possible distance pairs from Projects
This will create a complete distance matrix for all locations

Usage: python generate_distance_matrix.py
"""

import pandas as pd
from openpyxl import load_workbook
import sys
from itertools import combinations, permutations

try:
    sys.stdout.reconfigure(encoding='utf-8')
except: pass

print("=== Auto-Generate Distance Matrix ===\n")

# Load Config.xlsx
config_file = 'Config.xlsx'

# 1. Read Projects (locations)
df_projects = pd.read_excel(config_file, sheet_name='Projects')
projects = df_projects['Active'].dropna().tolist()

# 2. Read Home Location from Settings
df_settings = pd.read_excel(config_file, sheet_name='Settings')
home_location = df_settings[df_settings['Key'] == 'Home_Location']['Value'].values[0]

# 3. Combine all locations (Projects + Home)
all_locations = list(set(projects + [home_location]))
print(f"Found {len(all_locations)} unique locations:")
for loc in all_locations:
    print(f"  - {loc}")

# 4. Generate all possible pairs (A->B and B->A are different)
all_pairs = list(permutations(all_locations, 2))
print(f"\nTotal possible pairs: {len(all_pairs)}")

# 5. Load existing distances
wb = load_workbook(config_file)
ws = wb['Distances']

# Read existing pairs
existing_pairs = set()
for row in range(2, ws.max_row + 1):
    from_loc = ws.cell(row, 1).value
    to_loc = ws.cell(row, 2).value
    if from_loc and to_loc:
        existing_pairs.add((from_loc, to_loc))

print(f"Existing pairs in Distances: {len(existing_pairs)}")

# 6. Add missing pairs
added_count = 0
last_row = ws.max_row

for from_loc, to_loc in all_pairs:
    if (from_loc, to_loc) not in existing_pairs:
        last_row += 1
        ws.cell(last_row, 1, from_loc)
        ws.cell(last_row, 2, to_loc)
        ws.cell(last_row, 3, 0.0)  # Placeholder distance
        added_count += 1
        print(f"Added: {from_loc} -> {to_loc} (0.0 km)")

# 7. Save
if added_count > 0:
    wb.save(config_file)
    print(f"\n✓ Successfully added {added_count} new distance pairs")
    print(f"✓ Total pairs now: {last_row - 1}")
    print("\n⚠️  All new pairs are set to 0.0 km")
    print("   Please update them with actual distances in Config.xlsx")
else:
    print("\n✓ All distance pairs already exist!")

print("\n=== Done ===")
